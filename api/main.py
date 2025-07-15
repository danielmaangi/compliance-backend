from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import List, Dict, Any
import json
import io
import base64
import pandas as pd
from pypdf import PdfReader
from openpyxl import load_workbook
from docx import Document
import re

app = FastAPI(title="EO Compliance Analysis API")

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Keywords list
KEYWORDS = [
    'gender', 'transgender', 'transmen', 'transwomen', 'lgbtq', 'lgbt', 'dei',
    'diversity', 'equity', 'inclusion', 'gbv', 'trans-gender', 'trans-women',
    'trans-men', 'disparity', 'pregnant people', 'identity', 'inclusivity',
    'binary', 'non-binary', 'prejudice', 'pronouns', 'race', 'stereotype',
    'tgw', 'tg', 'transgender women', 'trans', 'protecting women', 'key pops',
    'key populations', 'mat', 'hormone', 'dreams', 'abortion', 'fsw',
    'female sex worker', 'food'
]

class FileData(BaseModel):
    filename: str
    content: str

class AnalyzeRequest(BaseModel):
    files: List[FileData]

def extract_sentences_from_text(text: str, keywords: List[str]) -> List[Dict[str, Any]]:
    """Extract sentences containing keywords from text."""
    sentences = re.split(r'[.!?]+', text)
    matches = []
    
    for sentence in sentences:
        sentence = sentence.strip()
        if len(sentence) < 10:  # Skip very short sentences
            continue
            
        for keyword in keywords:
            if keyword.lower() in sentence.lower():
                matches.append({
                    'keyword': keyword,
                    'sentence': sentence,
                    'exact_sentence': sentence
                })
                break  # Only match first keyword found
    
    return matches

def process_pdf(file_content: bytes, filename: str) -> List[Dict[str, Any]]:
    """Process PDF file and extract keyword matches."""
    try:
        pdf_reader = PdfReader(io.BytesIO(file_content))
        results = []
        
        for page_num, page in enumerate(pdf_reader.pages, 1):
            text = page.extract_text()
            matches = extract_sentences_from_text(text, KEYWORDS)
            
            for match in matches:
                results.append({
                    'file_path': filename,
                    'source_type': 'page',
                    'source_name': f'Page {page_num}',
                    'location': f'Page {page_num}',
                    'keyword': match['keyword'],
                    'exact_sentence': match['sentence'],
                    'partner': filename.split('.')[0]
                })
        
        return results
    except Exception as e:
        return [{'error': f'Error processing PDF {filename}: {str(e)}'}]

def process_excel(file_content: bytes, filename: str) -> List[Dict[str, Any]]:
    """Process Excel file and extract keyword matches."""
    try:
        workbook = load_workbook(io.BytesIO(file_content))
        results = []
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            for row_num, row in enumerate(sheet.iter_rows(values_only=True), 1):
                for col_num, cell_value in enumerate(row, 1):
                    if cell_value and isinstance(cell_value, str):
                        matches = extract_sentences_from_text(cell_value, KEYWORDS)
                        
                        for match in matches:
                            results.append({
                                'file_path': filename,
                                'source_type': 'worksheet',
                                'source_name': sheet_name,
                                'location': f'Row {row_num}, Column {col_num}',
                                'keyword': match['keyword'],
                                'exact_sentence': match['sentence'],
                                'partner': filename.split('.')[0]
                            })
        
        return results
    except Exception as e:
        return [{'error': f'Error processing Excel {filename}: {str(e)}'}]

def process_word(file_content: bytes, filename: str) -> List[Dict[str, Any]]:
    """Process Word document and extract keyword matches."""
    try:
        doc = Document(io.BytesIO(file_content))
        results = []
        
        for para_num, paragraph in enumerate(doc.paragraphs, 1):
            if paragraph.text.strip():
                matches = extract_sentences_from_text(paragraph.text, KEYWORDS)
                
                for match in matches:
                    results.append({
                        'file_path': filename,
                        'source_type': 'paragraph',
                        'source_name': f'Paragraph {para_num}',
                        'location': f'Paragraph {para_num}',
                        'keyword': match['keyword'],
                        'exact_sentence': match['sentence'],
                        'partner': filename.split('.')[0]
                    })
        
        return results
    except Exception as e:
        return [{'error': f'Error processing Word {filename}: {str(e)}'}]

@app.get("/")
async def root():
    return {"message": "EO Compliance Analysis API"}

@app.post("/analyze")
async def analyze_files(request: AnalyzeRequest):
    """Analyze uploaded files for compliance keywords."""
    try:
        if not request.files:
            raise HTTPException(status_code=400, detail="No files provided")
        
        all_results = []
        
        for file_data in request.files:
            filename = file_data.filename
            file_content = base64.b64decode(file_data.content)
            
            # Process based on file extension
            if filename.lower().endswith('.pdf'):
                results = process_pdf(file_content, filename)
            elif filename.lower().endswith('.xlsx') or filename.lower().endswith('.xls'):
                results = process_excel(file_content, filename)
            elif filename.lower().endswith('.docx'):
                results = process_word(file_content, filename)
            else:
                results = [{'error': f'Unsupported file type: {filename}'}]
            
            all_results.extend(results)
        
        # Filter out error results for statistics
        valid_results = [r for r in all_results if 'error' not in r]
        
        # Create summary statistics
        summary_stats = {
            'total_matches': len(valid_results),
            'files_processed': len(request.files),
            'keywords_found': len(set(result.get('keyword', '') for result in valid_results if 'keyword' in result)),
            'results': all_results
        }
        
        return summary_stats
    
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Internal server error: {str(e)}")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
