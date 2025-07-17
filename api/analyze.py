import json
import io
import base64
from typing import List, Dict, Any
import pandas as pd
from PyPDF2 import PdfReader
from openpyxl import load_workbook
from docx import Document
import re

# Keywords list
KEYWORDS = [
    'gender', 'transgender', 'transmen', 'transwomen', 'lgbtq', 'lgbt', ' dei ',
    'diversity', 'equity', 'inclusion', ' gbv ', 'trans-gender', 'trans-women',
    'trans-men', 'disparity', 'pregnant people', 'identity', 'inclusivity',
    'binary', 'non-binary', 'prejudice', 'pronouns', 'race', 'stereotype',
    ' tgw ', ' tg ', 'transgender women', ' trans ', 'protecting women', 'key pops',
    'key populations', ' mat ', 'hormone', ' dreams ', 'abortion', ' fsw',
    'female sex worker', 'food'
]

def extract_sentences_from_text(text: str, keywords: List[str]) -> List[Dict[str, Any]]:
    """Extract sentences containing keywords from text."""
    sentences = re.split(r'[.!?]+', text)
    matches = []
    
    for sentence in sentences:
        sentence = sentence.strip()
            
        for keyword in keywords:
            if keyword.lower() in sentence.lower():
                matches.append({
                    'keyword': keyword,
                    'sentence': sentence,
                    'exact_sentence': sentence
                })
                break  # Only match first keyword found
    
    return matches

def process_pdf(file_content: bytes, filename: str) -> List[Dict[str, Any]]:
    """Process PDF file and extract keyword matches."""
    try:
        pdf_reader = PdfReader(io.BytesIO(file_content))
        results = []
        
        for page_num, page in enumerate(pdf_reader.pages, 1):
            text = page.extract_text()
            matches = extract_sentences_from_text(text, KEYWORDS)
            
            for match in matches:
                results.append({
                    'file_path': filename,
                    'source_type': 'page',
                    'source_name': f'Page {page_num}',
                    'location': f'Page {page_num}',
                    'keyword': match['keyword'],
                    'exact_sentence': match['sentence'],
                    'partner': filename.split('.')[0]
                })
        
        return results
    except Exception as e:
        return [{'error': f'Error processing PDF {filename}: {str(e)}'}]

def process_excel(file_content: bytes, filename: str) -> List[Dict[str, Any]]:
    """Process Excel file and extract keyword matches."""
    try:
        workbook = load_workbook(io.BytesIO(file_content))
        results = []
        
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            for row_num, row in enumerate(sheet.iter_rows(values_only=True), 1):
                for col_num, cell_value in enumerate(row, 1):
                    if cell_value and isinstance(cell_value, str):
                        matches = extract_sentences_from_text(cell_value, KEYWORDS)
                        
                        for match in matches:
                            results.append({
                                'file_path': filename,
                                'source_type': 'worksheet',
                                'source_name': sheet_name,
                                'location': f'Row {row_num}, Column {col_num}',
                                'keyword': match['keyword'],
                                'exact_sentence': match['sentence'],
                                'partner': filename.split('.')[0]
                            })
        
        return results
    except Exception as e:
        return [{'error': f'Error processing Excel {filename}: {str(e)}'}]

def process_word(file_content: bytes, filename: str) -> List[Dict[str, Any]]:
    """Process Word document and extract keyword matches."""
    try:
        doc = Document(io.BytesIO(file_content))
        results = []
        
        for para_num, paragraph in enumerate(doc.paragraphs, 1):
            if paragraph.text.strip():
                matches = extract_sentences_from_text(paragraph.text, KEYWORDS)
                
                for match in matches:
                    results.append({
                        'file_path': filename,
                        'source_type': 'paragraph',
                        'source_name': f'Paragraph {para_num}',
                        'location': f'Paragraph {para_num}',
                        'keyword': match['keyword'],
                        'exact_sentence': match['sentence'],
                        'partner': filename.split('.')[0]
                    })
        
        return results
    except Exception as e:
        return [{'error': f'Error processing Word {filename}: {str(e)}'}]

def handler(request):
    """Main handler function for Vercel."""
    if request.method != 'POST':
        return {
            'statusCode': 405,
            'body': json.dumps({'error': 'Method not allowed'})
        }
    
    try:
        # Parse request body
        body = json.loads(request.body)
        files = body.get('files', [])
        
        if not files:
            return {
                'statusCode': 400,
                'body': json.dumps({'error': 'No files provided'})
            }
        
        all_results = []
        
        for file_data in files:
            filename = file_data['filename']
            file_content = base64.b64decode(file_data['content'])
            
            # Process based on file extension
            if filename.lower().endswith('.pdf'):
                results = process_pdf(file_content, filename)
            elif filename.lower().endswith('.xlsx'):
                results = process_excel(file_content, filename)
            elif filename.lower().endswith('.docx'):
                results = process_word(file_content, filename)
            else:
                results = [{'error': f'Unsupported file type: {filename}'}]
            
            all_results.extend(results)
        
        # Generate summary
        summary_df = pd.DataFrame(all_results)
        
        # Create summary statistics
        summary_stats = {
            'total_matches': len(all_results),
            'files_processed': len(files),
            'keywords_found': len(set(result.get('keyword', '') for result in all_results if 'keyword' in result)),
            'results': all_results
        }
        
        return {
            'statusCode': 200,
            'headers': {
                'Content-Type': 'application/json',
                'Access-Control-Allow-Origin': '*'
            },
            'body': json.dumps(summary_stats)
        }
    
    except Exception as e:
        return {
            'statusCode': 500,
            'body': json.dumps({'error': f'Internal server error: {str(e)}'})
        }